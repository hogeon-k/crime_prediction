from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook
import pandas as pd

from model.excel_model import ProcessResult, ValidationRule

_ENCODINGS = ("utf-8-sig", "euc-kr", "cp949")
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xls"}
MAX_UPLOAD_SIZE_BYTES = 20 * 1024 * 1024
MAX_UPLOAD_ROWS = 100_000
MAX_EXCEL_SHEETS = 1

# 업로드 파일 컬럼 별칭 → 표준 컬럼명
_COLUMN_ALIASES: dict[str, list[str]] = {
    "범죄_유형": [
        "범죄_유형",
        "범죄유형",
        "범죄 유형",
        "범죄중분류",
        "crime_type",
    ],
    "지역": ["지역", "region", "시도", "시도명"],
    "연도": ["연도", "날짜", "date", "year", "기준연도"],
    "발생_건수": [
        "발생_건수",
        "발생건수",
        "범죄발생건수",
        "범죄 발생 건수",
        "건수",
        "count",
        "발생 건수",
    ],
    "인구수": ["인구수", "인구", "population"],
}

# 인구 CSV 시도명(전체) → 범죄 CSV 시도명(약칭) 매핑
#    예: '서울특별시' → '서울', '강원특별자치도' → '강원'
_SIDO_SHORT: dict[str, str] = {
    "서울특별시": "서울",
    "부산광역시": "부산",
    "대구광역시": "대구",
    "인천광역시": "인천",
    "광주광역시": "광주",
    "대전광역시": "대전",
    "울산광역시": "울산",
    "세종특별자치시": "세종",
    "경기도": "경기",
    "강원도": "강원",
    "강원특별자치도": "강원",
    "충청북도": "충북",
    "충청남도": "충남",
    "전라북도": "전북",
    "전북특별자치도": "전북",
    "전라남도": "전남",
    "경상북도": "경북",
    "경상남도": "경남",
    "제주특별자치도": "제주",
}

_CRIME_SIDO_NAMES = tuple(_SIDO_SHORT.values())
_EXCLUDED_CRIME_REGION_COLUMNS = {"기타도시", "도시이외"}
_WIDE_POP_REGION_COLUMN = "행정구역별(읍면동)"
_WIDE_POP_ITEM_COLUMN = "항목"
_WIDE_POP_TOTAL_ITEM = "총인구[명]"


def _validate_upload_path(path: Path, allowed_roots: list[Path] | None = None) -> Path:
    """업로드 파일 경로가 읽을 수 있는 로컬 파일인지 확인한다."""
    resolved = path.expanduser().resolve()

    if not resolved.exists():
        raise ValueError(f"파일을 찾을 수 없습니다: {path}")
    if not resolved.is_file():
        raise ValueError(f"파일이 아닌 경로는 업로드할 수 없습니다: {path}")

    if allowed_roots:
        resolved_roots = [root.expanduser().resolve() for root in allowed_roots]
        if not any(resolved == root or root in resolved.parents for root in resolved_roots):
            raise ValueError(f"허용되지 않은 경로의 파일입니다: {path}")

    return resolved


def _validate_file_size(path: Path) -> None:
    size = path.stat().st_size
    if size > MAX_UPLOAD_SIZE_BYTES:
        limit_mb = MAX_UPLOAD_SIZE_BYTES / 1024 / 1024
        raise ValueError(f"파일 크기가 너무 큽니다: {size:,} bytes (최대 {limit_mb:.0f}MB)")


def _validate_extension(path: Path) -> str:
    suffix = path.suffix.lower()
    if suffix not in ALLOWED_EXTENSIONS:
        allowed = ", ".join(sorted(ALLOWED_EXTENSIONS))
        raise ValueError(f"지원하지 않는 파일 형식입니다: {suffix}\n지원 형식: {allowed}")
    return suffix


def _validate_row_limit(df: pd.DataFrame, file: str) -> None:
    if len(df) > MAX_UPLOAD_ROWS:
        raise ValueError(
            f"업로드 행 수가 너무 많습니다: {len(df):,}행 "
            f"(최대 {MAX_UPLOAD_ROWS:,}행): {file}"
        )


def _validate_excel_workbook(path: Path) -> None:
    if path.suffix.lower() != ".xlsx":
        return

    workbook = load_workbook(path, read_only=False, data_only=False)
    try:
        if len(workbook.sheetnames) > MAX_EXCEL_SHEETS:
            raise ValueError(
                f"Excel sheet 수가 너무 많습니다: {len(workbook.sheetnames)}개 "
                f"(최대 {MAX_EXCEL_SHEETS}개)"
            )

        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "f":
                        raise ValueError(
                            f"수식 셀은 업로드할 수 없습니다: "
                            f"{worksheet.title}!{cell.coordinate}"
                        )
    finally:
        workbook.close()


def _read_csv_safe(file: str) -> pd.DataFrame:
    """인코딩을 순차적으로 시도하며 CSV를 읽어 반환"""
    last_error: Exception = ValueError("알 수 없는 오류")
    for enc in _ENCODINGS:
        try:
            df = pd.read_csv(file, encoding=enc)
            _validate_row_limit(df, file)
            return df
        except UnicodeDecodeError as e:
            last_error = e
            continue
    raise ValueError(
        f"지원하는 인코딩으로 파일을 읽을 수 없습니다: {file}"
    ) from last_error


def _read_file_safe(file: str) -> pd.DataFrame:
    """CSV 또는 Excel 파일을 읽어 반환"""
    path = _validate_upload_path(Path(file))
    _validate_file_size(path)
    suffix = _validate_extension(path)

    if suffix in (".xlsx", ".xls"):
        _validate_excel_workbook(path)
        df = pd.read_excel(path, engine="openpyxl" if suffix == ".xlsx" else None)
        _validate_row_limit(df, file)
        return df
    if suffix == ".csv":
        return _read_csv_safe(str(path))

    raise ValueError(f"지원하지 않는 파일 형식입니다: {suffix}")


def _normalize_upload_columns(df: pd.DataFrame) -> pd.DataFrame:
    """업로드 파일 컬럼명을 표준 스키마로 통일"""
    df = df.copy()
    stripped = {col: str(col).strip() for col in df.columns}
    rename_map: dict[str, str] = {}

    for standard, aliases in _COLUMN_ALIASES.items():
        alias_set = {a.lower() for a in aliases}
        for col, name in stripped.items():
            if col in rename_map:
                continue
            if name in aliases or name.lower() in alias_set:
                rename_map[col] = standard

    df = df.rename(columns=rename_map)

    required = {"범죄_유형", "지역", "연도", "발생_건수", "인구수"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(
            f"필수 컬럼 누락: {sorted(missing)}\n"
            f"인식 가능한 컬럼 예시: {list(_COLUMN_ALIASES.keys())}"
        )

    if pd.api.types.is_datetime64_any_dtype(df["연도"]):
        df["연도"] = df["연도"].dt.year
    else:
        numeric = pd.to_numeric(df["연도"], errors="coerce")
        if numeric.notna().all() and numeric.between(1900, 2100).all():
            df["연도"] = numeric
        else:
            parsed = pd.to_datetime(df["연도"], errors="coerce")
            if parsed.notna().sum() > 0:
                df["연도"] = parsed.dt.year
            else:
                df["연도"] = numeric

    df["발생_건수"] = pd.to_numeric(df["발생_건수"], errors="coerce")
    if "인구수" in df.columns:
        df["인구수"] = pd.to_numeric(df["인구수"], errors="coerce")

    return df


def _extract_year(filename: str) -> int:
    """파일명에서 연도(20XX)를 정규식으로 추출"""
    match = re.search(r"(20\d{2})", Path(filename).stem)
    if match is None:
        raise ValueError(
            f"파일명에서 연도를 추출할 수 없습니다: {filename}\n"
            "파일명에 'YYYY' 형식의 연도(예: crime_2023.csv)가 포함되어야 합니다."
        )
    return int(match.group(1))


def _normalize_region(series: pd.Series) -> pd.Series:
    """지역명 정규화: 앞뒤 공백 제거, 연속 공백 단일화."""
    return series.astype(str).str.strip().str.replace(r"\s+", " ", regex=True)


def _normalize_sido(series: pd.Series) -> pd.Series:
    """시도 전체명/약칭/시군구 혼합명을 범죄 데이터의 시도 약칭으로 통일합니다."""
    normalized = _normalize_region(series)

    def to_sido(region: str) -> str:
        if region in _SIDO_SHORT:
            return _SIDO_SHORT[region]

        compact = region.replace(" ", "")
        for full_name, short_name in _SIDO_SHORT.items():
            if compact.startswith(full_name.replace(" ", "")):
                return short_name
        for short_name in _CRIME_SIDO_NAMES:
            if region == short_name or region.startswith(f"{short_name} "):
                return short_name
            if compact.startswith(short_name):
                return short_name
        return region

    return normalized.map(to_sido)


def _normalize_crime_region_to_sido(series: pd.Series) -> pd.Series:
    """범죄 파일의 시도/시군구 혼합 지역명을 시도 단위로 통일합니다."""
    return _normalize_sido(series)


def normalize_crime_type(value) -> str:
    """연도별 원천 파일의 공백/구분기호 차이를 같은 범죄유형명으로 통일합니다."""
    if pd.isna(value):
        return value
    return re.sub(r"[\s/·ㆍ]+", "", str(value).strip())


def _is_supported_crime_region_column(column: str) -> bool:
    normalized = str(column).strip().replace(" ", "")
    if normalized in _EXCLUDED_CRIME_REGION_COLUMNS:
        return False
    return not normalized.startswith("외국")


class CrimeService:

    def __init__(self) -> None:
        self._rule = ValidationRule()
        self.last_merge_report: dict[str, object] = {}

    def load_uploaded(self, file_path: str) -> ProcessResult:
        """표준 양식 Excel/CSV 단일 파일 업로드 로드"""
        try:
            raw = _read_file_safe(file_path)
            df = _normalize_upload_columns(raw)
            df["지역"] = _normalize_region(df["지역"])
            return ProcessResult(True, "업로드 로드 성공", df)
        except Exception as exc:
            return ProcessResult(False, f"업로드 로드 실패: {exc}")

    # 파일 로드 병합
    def load_and_merge(
        self,
        crime_files: list[str],
        pop_files: list[str],
    ) -> ProcessResult:
        try:
            crime_df = self._load_crime(crime_files)
            pop_df = self._load_population(pop_files)
            self.last_merge_report = self._build_merge_report(crime_df, pop_df)

            merged = pd.merge(
                crime_df,
                pop_df,
                on=["지역", "연도"],
                how="left",
            )
            missing_population = merged[merged["인구수"].isna()]
            if not missing_population.empty:
                examples = (
                    missing_population[["지역", "연도"]]
                    .drop_duplicates()
                    .head(10)
                    .to_dict(orient="records")
                )
                raise ValueError(
                    "공공데이터 병합 후 인구수 결측이 발생했습니다. "
                    f"범죄/인구 파일의 지역명 또는 연도를 확인하세요. 예시: {examples}"
                )
            merged.attrs["preprocessing_report"] = self.last_merge_report
            return ProcessResult(True, "병합 성공", merged)

        except Exception as exc:
            return ProcessResult(False, f"병합 실패: {exc}")

    def _load_crime(self, crime_files: list[str]) -> pd.DataFrame:
        """
        범죄 CSV를 읽어 시도 단위 long-format으로 변환합니다.

        연도별 집계 단위 차이 처리:
        - 2022: 시도/일부 시군구 혼합 컬럼 ('서울', '경기 고양', ...)
        - 2023: 시군구 단위 컬럼 ('서울종로구', '서울중구', ...)
        - 2024: 시군구 단위 컬럼 ('서울 종로구', '서울 중구', ...)
                → 시도 단위로 합산하여 연도 간 통일
        """
        frames: list[pd.DataFrame] = []

        for file in crime_files:
            year = _extract_year(file)
            crime = _read_file_safe(file)

            region_cols = [
                col
                for col in crime.columns
                if col not in ["범죄대분류", "범죄중분류"]
                and _is_supported_crime_region_column(col)
            ]

            crime = crime.melt(
                id_vars=["범죄대분류", "범죄중분류"],
                value_vars=region_cols,
                var_name="지역",
                value_name="발생_건수",
            )

            crime = crime[~crime["지역"].str.startswith("외국")].copy()
            crime["범죄_유형"] = crime["범죄중분류"].map(normalize_crime_type)
            crime["연도"] = year
            crime["지역"] = _normalize_crime_region_to_sido(crime["지역"])
            crime["발생_건수"] = pd.to_numeric(crime["발생_건수"], errors="coerce")

            crime = pd.DataFrame(
                crime.groupby(["범죄_유형", "지역", "연도"], as_index=False)[
                    "발생_건수"
                ].sum()
            )

            frames.append(crime[["범죄_유형", "지역", "연도", "발생_건수"]])

        return pd.concat(frames, ignore_index=True)

    def _load_population(self, pop_files: list[str]) -> pd.DataFrame:
        """
        인구 CSV를 읽어 시도·연도 단위로 집계합니다.

        연도별 컬럼명 차이 처리:
        - 2022·2023: '통계년월'
        - 2024     : '기준연월'
        """
        frames: list[pd.DataFrame] = []

        for file in pop_files:
            raw = _read_file_safe(file)

            if {_WIDE_POP_REGION_COLUMN, _WIDE_POP_ITEM_COLUMN}.issubset(raw.columns):
                pop = self._load_wide_population(raw)
                frames.append(pop)
                continue

            required = {"시도명", "계"}
            missing = required - set(raw.columns)
            if missing:
                raise ValueError(
                    f"지원하지 않는 인구 파일 구조입니다: {file}. "
                    f"누락 컬럼: {sorted(missing)}"
                )

            sido_short = _normalize_sido(raw["시도명"])

            연월_col = "기준연월" if "기준연월" in raw.columns else "통계년월"
            if 연월_col not in raw.columns:
                raise ValueError(
                    f"인구 파일에서 연월 컬럼을 찾을 수 없습니다: {file}. "
                    "기대 컬럼: 기준연월 또는 통계년월"
                )
            연도 = pd.to_numeric(
                raw[연월_col].astype(str).str[:4], errors="coerce"
            ).astype("Int64")

            인구수 = pd.to_numeric(raw["계"], errors="coerce")

            pop = pd.concat(
                [
                    sido_short.rename("지역"),
                    연도.rename("연도"),
                    인구수.rename("인구수"),
                ],
                axis=1,
            )

            pop = pd.DataFrame(
                pop.groupby(["지역", "연도"], as_index=False)["인구수"].sum()
            )
            frames.append(pop)

        combined = pd.concat(frames, ignore_index=True)
        combined = combined[combined["지역"].isin(_CRIME_SIDO_NAMES)].copy()
        combined["인구수"] = pd.to_numeric(combined["인구수"], errors="coerce")
        combined = combined.dropna(subset=["지역", "연도", "인구수"])
        combined["연도"] = combined["연도"].astype("Int64")
        return pd.DataFrame(
            combined.groupby(["지역", "연도"], as_index=False)["인구수"].mean()
        )

    def _load_wide_population(self, raw: pd.DataFrame) -> pd.DataFrame:
        total = raw[raw[_WIDE_POP_ITEM_COLUMN].astype(str).str.strip() == _WIDE_POP_TOTAL_ITEM].copy()
        if total.empty:
            raise ValueError(f"wide 인구 파일에 '{_WIDE_POP_TOTAL_ITEM}' 항목이 없습니다.")

        year_columns = [
            column
            for column in total.columns
            if re.fullmatch(r"\s*20\d{2}\s*년\s*", str(column))
        ]
        if not year_columns:
            raise ValueError("wide 인구 파일에서 'YYYY 년' 형식의 연도 컬럼을 찾을 수 없습니다.")

        pop = total.melt(
            id_vars=[_WIDE_POP_REGION_COLUMN],
            value_vars=year_columns,
            var_name="연도",
            value_name="인구수",
        )
        pop["지역"] = _normalize_sido(pop[_WIDE_POP_REGION_COLUMN])
        pop["연도"] = pd.to_numeric(
            pop["연도"].astype(str).str.extract(r"(20\d{2})")[0],
            errors="coerce",
        ).astype("Int64")
        pop["인구수"] = pd.to_numeric(pop["인구수"], errors="coerce")
        pop = pop[pop["지역"].isin(_CRIME_SIDO_NAMES)].copy()
        return pop[["지역", "연도", "인구수"]]

    def _build_merge_report(
        self,
        crime_df: pd.DataFrame,
        pop_df: pd.DataFrame,
    ) -> dict[str, object]:
        crime_keys = crime_df[["지역", "연도"]].drop_duplicates()
        pop_keys = pop_df[["지역", "연도"]].drop_duplicates()
        missing_keys = crime_keys.merge(
            pop_keys,
            on=["지역", "연도"],
            how="left",
            indicator=True,
        )
        missing_keys = missing_keys[missing_keys["_merge"] == "left_only"][["지역", "연도"]]

        return {
            "crime_year_counts": {
                int(year): int(count)
                for year, count in crime_df.groupby("연도").size().sort_index().items()
            },
            "population_year_counts": {
                int(year): int(count)
                for year, count in pop_keys.groupby("연도").size().sort_index().items()
            },
            "crime_regions": sorted(crime_df["지역"].dropna().astype(str).unique().tolist()),
            "population_regions": sorted(pop_df["지역"].dropna().astype(str).unique().tolist()),
            "merge_failed_key_count": int(len(missing_keys)),
            "merge_failed_keys": missing_keys.head(20).to_dict(orient="records"),
            "crime_missing_values": {
                column: int(value)
                for column, value in crime_df.isna().sum().items()
            },
            "population_missing_values": {
                column: int(value)
                for column, value in pop_df.isna().sum().items()
            },
        }

    # 컬럼 검증 연도 병합
    def validate(self, df: pd.DataFrame) -> ProcessResult:
        missing = self._rule.required_columns - set(df.columns)
        if missing:
            return ProcessResult(False, f"컬럼 누락: {missing}")

        if "연도" in df.columns:
            valid_set = set(self._rule.valid_years)
            actual_years = set(df["연도"].dropna().unique())
            invalid_years = actual_years - valid_set
            if invalid_years:
                return ProcessResult(
                    False,
                    f"유효하지 않은 연도 값 포함: {sorted(invalid_years)}\n"
                    f"허용 범위: {min(valid_set)}~{max(valid_set)}",
                )

        return ProcessResult(True, "검증 성공", df)

    # 결측치 처리
    def handle_missing(self, df: pd.DataFrame) -> ProcessResult:
        df = df.copy()

        df.dropna(subset=["범죄_유형", "지역", "연도"], inplace=True)
        df["발생_건수"] = df["발생_건수"].fillna(0)

        df["인구수"] = df.groupby("지역")["인구수"].transform(
            lambda x: x.fillna(x.mean())
        )
        global_mean = df["인구수"].mean()
        df["인구수"] = df["인구수"].fillna(global_mean if pd.notna(global_mean) else 0)

        return ProcessResult(True, "결측치 처리 완료", df)

    # 타입 변환, 범죄율 계산
    def convert_types(self, df: pd.DataFrame) -> ProcessResult:
        try:
            df = df.copy()

            df["연도"] = pd.to_numeric(df["연도"], errors="coerce").astype("Int64")
            df["발생_건수"] = (
                pd.to_numeric(df["발생_건수"], errors="coerce").fillna(0).astype(int)
            )
            df["인구수"] = (
                pd.to_numeric(df["인구수"], errors="coerce").fillna(0).astype(int)
            )

            df["범죄율"] = (
                df["발생_건수"] / df["인구수"].replace(0, pd.NA) * 100_000
            ).fillna(0.0)

            return ProcessResult(True, "타입 변환 성공", df)

        except Exception as exc:
            return ProcessResult(False, f"변환 실패: {exc}")
